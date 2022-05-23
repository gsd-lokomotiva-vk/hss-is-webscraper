from openpyxl import Workbook, load_workbook
from typing import List


workbook_extension = ".xlsx"


def create_new_workbook(name: str, path: str):
    """Returns newly created workbook"""
    wb = Workbook()
    wb.save(path + name + workbook_extension)


def append_rows_to_existing_workbook(filename: str, filepath: str, rows = List[List]):
    wb = load_workbook(filepath + filename + workbook_extension)
    ws = wb.active
    if isinstance(rows, list):
        try:
            if isinstance(rows[0], list):
                for row in rows:
                    ws.append(row)
            else:
                ws.append(rows)
        except IndexError:
            pass
    
    wb.save(filepath + filename + workbook_extension)


def export_to_excel_workbook(sheet_name: str, data: List[List], filepath: str, filename: str, headers: List = None):
    if headers is not None:
        if not check_correct_header_count(data[0], headers):
            return -1
    wb = Workbook()
    ws = wb.active
    extension = ".xlsx"
    ws.title = sheet_name

    row = 1
    column = 1

    for header in headers:
        ws.cell(row=row, column=column, value=header)
        column += 1
    
    row += 1
    column = 1

    for row_data in data:
        for cell_value in row_data:
            ws.cell(row=row, column=column, value=cell_value)
            column += 1
        row += 1
        column = 1
    
    wb.save(filepath + filename + extension)
    
    
def check_correct_header_count(row_data: List, headers: List):
    if len(row_data) == len(headers):
        return True
    return False
